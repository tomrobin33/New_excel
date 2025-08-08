import logging
import os
from typing import Any, List, Dict, Optional

from mcp.server.fastmcp import FastMCP

# Import exceptions
from excel_mcp.exceptions import (
    ValidationError,
    WorkbookError,
    SheetError,
    DataError,
    FormattingError,
    CalculationError,
    PivotError,
    ChartError
)

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
    validate_range_in_sheet_operation as validate_range_impl
)
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.workbook import get_workbook_info
from excel_mcp.data import write_data
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.sheet import (
    copy_sheet,
    delete_sheet,
    rename_sheet,
    merge_range,
    unmerge_range,
    get_merged_ranges,
)
import requests
import uuid
from fastapi import FastAPI, Form
from openpyxl import load_workbook
import paramiko
import shutil
import tempfile

app = FastAPI()
TEMP_DIR = "/tmp"

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "excel-mcp.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# Initialize FastMCP server
mcp = FastMCP(
    "excel-mcp",
    version="0.1.5",
    description="Excel MCP Server for manipulating Excel files",
    dependencies=["openpyxl>=3.1.5"],
    env_vars={
        "EXCEL_FILES_PATH": {
            "description": "Path to Excel files directory",
            "required": False,
            "default": EXCEL_FILES_PATH
        }
    }
)

def get_excel_path(filename: str) -> str:
    """Get full path to Excel file.
    支持绝对路径和相对路径（如tmp_files/xxx.xlsx）。
    """
    # If filename is already an absolute path, return it
    if os.path.isabs(filename):
        return filename

    # Check if in SSE mode (EXCEL_FILES_PATH is not None)
    if EXCEL_FILES_PATH is None:
        # 支持相对路径，返回其绝对路径
        return os.path.abspath(filename)

    # In SSE mode, if it's a relative path, resolve it based on EXCEL_FILES_PATH
    return os.path.join(EXCEL_FILES_PATH, filename)

@mcp.tool()
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply Excel formula to cell.
    Excel formula will write to cell with verification.
    """
    try:
        full_path = get_excel_path(filepath)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"
            
        # If valid, apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl
        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise

@mcp.tool()
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it."""
    try:
        full_path = get_excel_path(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise

@mcp.tool()
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None
) -> str:
    """Apply formatting to a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.formatting import format_range as format_range_func
        
        # Convert None values to appropriate defaults for the underlying function
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,  # This can be None
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,  # This can be None
            font_color=font_color,  # This can be None
            bg_color=bg_color,  # This can be None
            border_style=border_style,  # This can be None
            border_color=border_color,  # This can be None
            number_format=number_format,  # This can be None
            alignment=alignment,  # This can be None
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,  # This can be None
            conditional_format=conditional_format  # This can be None
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise

@mcp.tool()
def read_data_from_excel(
    filepath: str,
    sheet_name: Optional[str] = None,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False,
    max_rows: int = 100,  # 添加行数限制
    max_cells: int = 1000  # 添加单元格数量限制
) -> str:
    """
    只支持通过URL读取Excel文件，简化逻辑，避免大模型错误思考。
    添加了数据大小限制，防止返回过大的数据块。
    """
    import requests
    import uuid
    temp_file = None
    try:
        # 只允许URL输入
        if not (filepath.startswith("http://") or filepath.startswith("https://")):
            return "Error: 只支持通过URL读取Excel文件，请输入有效的http/https链接。"
        
        # 验证参数
        if max_rows <= 0 or max_cells <= 0:
            return "Error: max_rows 和 max_cells 必须大于0"
        
        temp_file = f"/tmp/{uuid.uuid4()}.xlsx"
        
        # 添加请求头和超时设置
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        r = requests.get(filepath, stream=True, headers=headers, timeout=30)
        
        # 检查HTTP状态码
        if r.status_code != 200:
            return f"Error: 无法下载文件，HTTP状态码: {r.status_code}"
        
        # 下载文件
        file_size = 0
        with open(temp_file, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:  # 过滤掉keep-alive新块
                    f.write(chunk)
                    file_size += len(chunk)
        
        # 检查文件大小
        if file_size == 0:
            return "Error: 下载的文件为空"
        
        # 验证文件是否为有效的Excel文件
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_file, read_only=True)
            wb.close()
        except Exception as excel_error:
            return f"Error: 文件不是有效的Excel文件 - {str(excel_error)}"
        
        full_path = temp_file
        from excel_mcp.data import read_excel_range_with_metadata
        result = read_excel_range_with_metadata(
            full_path, 
            sheet_name, 
            start_cell, 
            end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"
        
        # 限制返回的数据大小
        cells = result.get("cells", [])
        if len(cells) > max_cells:
            cells = cells[:max_cells]
            result["cells"] = cells
            result["truncated"] = True
            result["total_cells"] = len(cells)
            result["max_cells_limit"] = max_cells
        
        # 限制行数
        if len(cells) > max_rows * 10:  # 假设每行最多10列
            cells = cells[:max_rows * 10]
            result["cells"] = cells
            result["truncated"] = True
            result["max_rows_limit"] = max_rows
        
        import json
        json_result = json.dumps(result, indent=2, default=str)
        
        # 检查JSON大小
        if len(json_result) > 50000:  # 50KB限制
            # 返回简化版本
            simplified_result = {
                "range": result.get("range", ""),
                "sheet_name": result.get("sheet_name", ""),
                "total_cells": len(cells),
                "preview_cells": cells[:10],  # 只返回前10个单元格作为预览
                "message": f"数据过大，已截断。总共{len(cells)}个单元格，只显示前10个作为预览。"
            }
            return json.dumps(simplified_result, indent=2, default=str)
        
        return json_result
    except requests.exceptions.RequestException as e:
        return f"Error: 网络请求失败 - {str(e)}"
    except Exception as e:
        logger.error(f"Error reading data: {e}")
        return f"Error: {e}"
    finally:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)

@mcp.tool()
def preview_excel_data(
    filepath: str,
    sheet_name: Optional[str] = None,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    max_preview_rows: int = 5,
    max_preview_cols: int = 5
) -> str:
    """
    预览Excel文件数据，返回小规模的数据样本。
    适用于快速查看文件结构和内容，避免返回过大的数据块。
    """
    import requests
    import uuid
    temp_file = None
    try:
        # 只允许URL输入
        if not (filepath.startswith("http://") or filepath.startswith("https://")):
            return "Error: 只支持通过URL读取Excel文件，请输入有效的http/https链接。"
        
        temp_file = f"/tmp/{uuid.uuid4()}.xlsx"
        
        # 添加请求头和超时设置
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        r = requests.get(filepath, stream=True, headers=headers, timeout=30)
        
        # 检查HTTP状态码
        if r.status_code != 200:
            return f"Error: 无法下载文件，HTTP状态码: {r.status_code}"
        
        # 下载文件
        file_size = 0
        with open(temp_file, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:  # 过滤掉keep-alive新块
                    f.write(chunk)
                    file_size += len(chunk)
        
        # 检查文件大小
        if file_size == 0:
            return "Error: 下载的文件为空"
        
        # 验证文件是否为有效的Excel文件
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_file, read_only=True)
        except Exception as excel_error:
            return f"Error: 文件不是有效的Excel文件 - {str(excel_error)}"
        
        # 自动适应sheet_name
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
            
        ws = wb[sheet_name]
        
        # 获取工作表信息
        sheet_info = {
            "sheet_name": sheet_name,
            "total_rows": ws.max_row,
            "total_columns": ws.max_column,
            "preview_data": []
        }
        
        # 生成预览数据
        preview_rows = min(max_preview_rows, ws.max_row)
        preview_cols = min(max_preview_cols, ws.max_column)
        
        for row in range(1, preview_rows + 1):
            row_data = []
            for col in range(1, preview_cols + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            sheet_info["preview_data"].append(row_data)
        
        wb.close()
        
        import json
        return json.dumps(sheet_info, indent=2, default=str)
        
    except requests.exceptions.RequestException as e:
        return f"Error: 网络请求失败 - {str(e)}"
    except Exception as e:
        logger.error(f"Error previewing data: {e}")
        return f"Error: {e}"
    finally:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)

@mcp.tool()
def create_workbook(filepath: str) -> str:
    """Create new Excel workbook. 允许先创建空表，后续写入时自动保存/上传。"""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_workbook as create_workbook_impl
        create_workbook_impl(full_path)
        return f"Created workbook at {full_path} (空表，待写入数据后自动上传)"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise

@mcp.tool()
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """Create new worksheet in workbook. 允许先创建空表，后续写入时自动保存/上传。"""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_sheet as create_worksheet_impl
        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"] + " (空表，待写入数据后自动上传)"
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise

@mcp.tool()
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = ""
) -> str:
    """Create chart in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise

@mcp.tool()
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean"
) -> str:
    """Create pivot table in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise

@mcp.tool()
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9"
) -> str:
    """Creates a native Excel table from a specified range of data."""
    try:
        full_path = get_excel_path(filepath)
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style
        )
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise

@mcp.tool()
def copy_worksheet(
    filepath: str,
    source_sheet: str,
    target_sheet: str
) -> str:
    """Copy worksheet within workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise

@mcp.tool()
def delete_worksheet(
    filepath: str,
    sheet_name: str
) -> str:
    """Delete worksheet from workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise

@mcp.tool()
def rename_worksheet(
    filepath: str,
    old_name: str,
    new_name: str
) -> str:
    """Rename worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise

@mcp.tool()
def get_workbook_metadata(
    filepath: str,
    include_ranges: bool = False,
    max_sheets_info: int = 10  # 添加工作表信息限制
) -> str:
    """只支持通过URL读取Excel文件元数据，简化逻辑。"""
    import requests, uuid, os
    temp_file = None
    try:
        if not (filepath.startswith("http://") or filepath.startswith("https://")):
            return "Error: 只支持通过URL读取Excel文件元数据，请输入有效的http/https链接。"
        
        temp_file = f"/tmp/{uuid.uuid4()}.xlsx"
        
        # 添加请求头和超时设置
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        r = requests.get(filepath, stream=True, headers=headers, timeout=30)
        
        # 检查HTTP状态码
        if r.status_code != 200:
            return f"Error: 无法下载文件，HTTP状态码: {r.status_code}"
        
        # 检查Content-Type
        content_type = r.headers.get('content-type', '').lower()
        if 'excel' not in content_type and 'spreadsheet' not in content_type and 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' not in content_type:
            logger.warning(f"Content-Type不是Excel文件: {content_type}")
        
        # 下载文件
        file_size = 0
        with open(temp_file, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:  # 过滤掉keep-alive新块
                    f.write(chunk)
                    file_size += len(chunk)
        
        # 检查文件大小
        if file_size == 0:
            return "Error: 下载的文件为空"
        
        # 验证文件是否为有效的Excel文件
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_file, read_only=True)
            wb.close()
        except Exception as excel_error:
            return f"Error: 文件不是有效的Excel文件 - {str(excel_error)}"
        
        full_path = temp_file
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        
        # 限制返回的工作表信息数量
        if isinstance(result, dict) and "sheets" in result:
            sheets = result["sheets"]
            if len(sheets) > max_sheets_info:
                result["sheets"] = sheets[:max_sheets_info]
                result["truncated"] = True
                result["total_sheets"] = len(sheets)
                result["max_sheets_limit"] = max_sheets_info
        
        result_str = str(result)
        
        # 检查结果大小
        if len(result_str) > 10000:  # 10KB限制
            simplified_result = {
                "message": "元数据过大，已截断",
                "total_size": len(result_str),
                "preview": result_str[:1000] + "..." if len(result_str) > 1000 else result_str
            }
            return str(simplified_result)
        
        return result_str
    except requests.exceptions.RequestException as e:
        return f"Error: 网络请求失败 - {str(e)}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        return f"Error: {e}"
    finally:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)

@mcp.tool()
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Merge a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise

@mcp.tool()
def unmerge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Unmerge a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise

@mcp.tool()
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """Get merged cells in a worksheet."""
    try:
        full_path = get_excel_path(filepath)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise

@mcp.tool()
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> str:
    """Copy a range of cells to another location."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import copy_range_operation
        result = copy_range_operation(
            full_path,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name  # Use source sheet if target_sheet is None
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise

@mcp.tool()
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up"
) -> str:
    """Delete a range of cells and shift remaining cells."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import delete_range_operation
        result = delete_range_operation(
            full_path,
            sheet_name,
            start_cell,
            end_cell,
            shift_direction
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise

@mcp.tool()
def validate_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> str:
    """Validate if a range exists and is properly formatted."""
    try:
        full_path = get_excel_path(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise

@mcp.tool()
def get_data_validation_info(
    filepath: str,
    sheet_name: str
) -> str:
    """
    Get all data validation rules in a worksheet.
    
    This tool helps identify which cell ranges have validation rules
    and what types of validation are applied.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        
    Returns:
        JSON string containing all validation rules in the worksheet
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from excel_mcp.cell_validation import get_all_validation_ranges
        
        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
            
        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()
        
        if not validations:
            return "No data validation rules found in this worksheet"
            
        import json
        return json.dumps({
            "sheet_name": sheet_name,
            "validation_rules": validations
        }, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise

@mcp.tool()
def write_data_to_excel(
    filepath: str,
    sheet_name: Optional[str] = None,
    data: Optional[List[List]] = None,
    start_cell: str = "A1",
) -> str:
    """
    【用途说明】
    批量将结构化数据（如JSON、列表的列表）写入指定Excel文件和工作表。
    写入数据后会自动上传Excel文件到服务器（SFTP 8.154.74.79），并返回公网下载链接。
    适用于将PPT、表格、文本等结构化内容自动导出为Excel并获取公网访问。

    【参数说明】
    - filepath: Excel文件路径（如 "output.xlsx"）
    - sheet_name: 工作表名称（如 "Sheet1"，可选，默认第一个sheet）
    - data: 二维数组，每个子数组为一行。例如：[[1, "内容1"], [1, "内容2"], [2, "内容3"]]
    - start_cell: 写入起始单元格，通常为 "A1"

    【功能说明】
    - 写入数据后，自动将Excel文件上传到服务器（8.154.74.79:/root/files/），并生成公网下载链接（http://8.154.74.79:8001/文件名）。
    - 返回值中包含“公网下载链接”。

    【推荐用法】
    1. 先用 create_workbook 创建Excel文件（可选）
    2. 处理JSON等数据为二维数组
    3. 调用 write_data_to_excel 一步写入所有内容
    4. 系统会自动保存、上传，并返回公网下载链接

    【代码示例】
    >>> data = [[1, "标题1"], [1, "内容1"], [2, "标题2"], [2, "内容2"]]
    >>> write_data_to_excel(filepath="output.xlsx", sheet_name="Sheet1", data=data, start_cell="A1")
    # 返回示例：
    # Data written to Sheet1
    # 公网下载链接: http://8.154.74.79:8001/xxxx.xlsx

    【注意】
    - 这是唯一推荐的批量写入并上传数据到Excel的工具！
    - 不要用 apply_formula/format_range 写文本内容。
    """
    try:
        full_path = get_excel_path(filepath)
        # 如果文件不存在，自动创建
        if not os.path.exists(full_path):
            from excel_mcp.workbook import create_workbook as create_workbook_impl
            create_workbook_impl(full_path)
        # 如果sheet不存在，自动创建
        from openpyxl import load_workbook
        wb = load_workbook(full_path)
        if sheet_name and sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            wb.save(full_path)
            wb.close()
        # 写入数据
        result = write_data(full_path, sheet_name, data, start_cell)
        # 自动上传到服务器
        processed_filename = f"uploaded_{uuid.uuid4().hex}.xlsx"
        processed_path = os.path.join("/tmp", processed_filename)
        shutil.copy(full_path, processed_path)
        remote_path = f"/root/files/{processed_filename}"
        transport = paramiko.Transport(("8.156.74.79", 22))
        transport.connect(username="root", password="zfsZBC123.")
        sftp = paramiko.SFTPClient.from_transport(transport)
        if sftp is not None:
            sftp.put(processed_path, remote_path)
            sftp.close()
        if transport is not None:
            transport.close()
        download_url = f"http://8.156.74.79:8001/{processed_filename}"
        return f"{result['message']}\n公网下载链接: {download_url}"
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise

@mcp.tool()
def read_excel_data_in_batches(
    filepath: str,
    sheet_name: Optional[str] = None,
    batch_size: int = 50,
    start_row: int = 1,
    end_row: Optional[int] = None,
    columns: Optional[List[str]] = None
) -> str:
    """
    分批读取Excel文件数据，避免一次性读取大量数据导致"Chunk too big"错误。
    
    【参数说明】
    - filepath: Excel文件URL
    - sheet_name: 工作表名称（可选）
    - batch_size: 每批读取的行数（默认50行）
    - start_row: 开始读取的行号（默认1）
    - end_row: 结束读取的行号（可选，默认读取到文件末尾）
    - columns: 要读取的列（可选，默认读取所有列）
    
    【返回值】
    返回当前批次的数据和下一批次的读取信息，便于大模型进行分批处理。
    
    【使用示例】
    1. 第一次调用：read_excel_data_in_batches(filepath="url", batch_size=50)
    2. 根据返回的next_batch_info继续读取下一批
    3. 重复直到读取完所有数据
    """
    import requests
    import uuid
    temp_file = None
    try:
        # 只允许URL输入
        if not (filepath.startswith("http://") or filepath.startswith("https://")):
            return "Error: 只支持通过URL读取Excel文件，请输入有效的http/https链接。"
        
        # 验证参数
        if batch_size <= 0:
            return "Error: batch_size 必须大于0"
        if start_row <= 0:
            return "Error: start_row 必须大于0"
        
        temp_file = f"/tmp/{uuid.uuid4()}.xlsx"
        
        # 添加请求头和超时设置
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        r = requests.get(filepath, stream=True, headers=headers, timeout=30)
        
        # 检查HTTP状态码
        if r.status_code != 200:
            return f"Error: 无法下载文件，HTTP状态码: {r.status_code}"
        
        # 下载文件
        file_size = 0
        with open(temp_file, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:  # 过滤掉keep-alive新块
                    f.write(chunk)
                    file_size += len(chunk)
        
        # 检查文件大小
        if file_size == 0:
            return "Error: 下载的文件为空"
        
        # 验证文件是否为有效的Excel文件
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_file, read_only=True)
        except Exception as excel_error:
            return f"Error: 文件不是有效的Excel文件 - {str(excel_error)}"
        
        # 自动适应sheet_name
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
            
        ws = wb[sheet_name]
        
        # 确定读取范围
        max_row = ws.max_row
        if end_row is None:
            end_row = max_row
        else:
            end_row = min(end_row, max_row)
        
        # 计算当前批次的范围
        current_end_row = min(start_row + batch_size - 1, end_row)
        
        # 读取当前批次数据
        batch_data = []
        for row in range(start_row, current_end_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            batch_data.append(row_data)
        
        wb.close()
        
        # 构建返回结果
        result = {
            "filepath": filepath,
            "sheet_name": sheet_name,
            "batch_info": {
                "current_batch": {
                    "start_row": start_row,
                    "end_row": current_end_row,
                    "rows_count": len(batch_data)
                },
                "total_progress": {
                    "total_rows": end_row - start_row + 1,
                    "read_rows": current_end_row - start_row + 1,
                    "remaining_rows": max(0, end_row - current_end_row)
                }
            },
            "data": batch_data
        }
        
        # 如果还有更多数据，提供下一批次的读取信息
        if current_end_row < end_row:
            result["next_batch_info"] = {
                "start_row": current_end_row + 1,
                "end_row": end_row,
                "remaining_batches": (end_row - current_end_row) // batch_size + 1
            }
            result["message"] = f"已读取第{start_row}到{current_end_row}行，共{len(batch_data)}行数据。还有{result['next_batch_info']['remaining_batches']}批数据待读取。"
        else:
            result["message"] = f"已读取完所有数据，共{len(batch_data)}行。"
        
        import json
        return json.dumps(result, indent=2, default=str)
        
    except requests.exceptions.RequestException as e:
        return f"Error: 网络请求失败 - {str(e)}"
    except Exception as e:
        logger.error(f"Error reading data in batches: {e}")
        return f"Error: {e}"
    finally:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)

@mcp.tool()
def get_excel_file_info(
    filepath: str,
    sheet_name: Optional[str] = None
) -> str:
    """
    获取Excel文件的基本信息，帮助制定分批读取策略。
    
    【返回信息】
    - 文件大小
    - 工作表数量
    - 指定工作表的总行数和列数
    - 建议的批次大小
    - 预估的总批次数
    
    【用途】
    在开始分批读取前，先获取文件信息，制定合适的读取策略。
    """
    import requests
    import uuid
    temp_file = None
    try:
        # 只允许URL输入
        if not (filepath.startswith("http://") or filepath.startswith("https://")):
            return "Error: 只支持通过URL读取Excel文件，请输入有效的http/https链接。"
        
        temp_file = f"/tmp/{uuid.uuid4()}.xlsx"
        
        # 添加请求头和超时设置
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        r = requests.get(filepath, stream=True, headers=headers, timeout=30)
        
        # 检查HTTP状态码
        if r.status_code != 200:
            return f"Error: 无法下载文件，HTTP状态码: {r.status_code}"
        
        # 下载文件
        file_size = 0
        with open(temp_file, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:  # 过滤掉keep-alive新块
                    f.write(chunk)
                    file_size += len(chunk)
        
        # 检查文件大小
        if file_size == 0:
            return "Error: 下载的文件为空"
        
        # 验证文件是否为有效的Excel文件
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_file, read_only=True)
        except Exception as excel_error:
            return f"Error: 文件不是有效的Excel文件 - {str(excel_error)}"
        
        # 获取所有工作表信息
        sheets_info = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            sheets_info.append({
                "name": sheet,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "estimated_size_kb": (ws.max_row * ws.max_column * 50) // 1024  # 粗略估算
            })
        
        # 获取指定工作表信息
        target_sheet_info = None
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            target_sheet_info = {
                "name": sheet_name,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "estimated_size_kb": (ws.max_row * ws.max_column * 50) // 1024
            }
        
        wb.close()
        
        # 计算建议的批次大小
        if target_sheet_info:
            total_cells = target_sheet_info["rows"] * target_sheet_info["columns"]
            if total_cells > 10000:
                suggested_batch_size = 20
            elif total_cells > 5000:
                suggested_batch_size = 50
            elif total_cells > 1000:
                suggested_batch_size = 100
            else:
                suggested_batch_size = 200
            
            estimated_batches = (target_sheet_info["rows"] + suggested_batch_size - 1) // suggested_batch_size
        else:
            suggested_batch_size = 50
            estimated_batches = 0
        
        result = {
            "file_info": {
                "filepath": filepath,
                "file_size_bytes": file_size,
                "file_size_mb": round(file_size / (1024 * 1024), 2)
            },
            "workbook_info": {
                "total_sheets": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "sheets_info": sheets_info
            },
            "target_sheet_info": target_sheet_info,
            "reading_strategy": {
                "suggested_batch_size": suggested_batch_size,
                "estimated_batches": estimated_batches,
                "recommendation": f"建议使用{suggested_batch_size}行作为批次大小，预计需要{estimated_batches}次读取完成。"
            }
        }
        
        import json
        return json.dumps(result, indent=2, default=str)
        
    except requests.exceptions.RequestException as e:
        return f"Error: 网络请求失败 - {str(e)}"
    except Exception as e:
        logger.error(f"Error getting file info: {e}")
        return f"Error: {e}"
    finally:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)

@mcp.tool()
def extract_tables_from_document(
    document_url: str,
    output_filename: str = None,
    auto_upload: bool = True
) -> str:
    """
    从PPT、Word、PDF文档中提取表格数据并转换为Excel格式。
    
    【功能说明】
    - 支持从PPT(.pptx/.ppt)、Word(.docx/.doc)、PDF(.pdf)文件中提取表格
    - 自动识别文档中的表格结构
    - 将提取的表格保存为Excel文件，每个表格创建一个工作表
    - 可选择自动上传到服务器并返回公网下载链接
    
    【参数说明】
    - document_url: 文档的URL地址（必须是以http://或https://开头的有效链接）
    - output_filename: 输出的Excel文件名（可选，默认自动生成）
    - auto_upload: 是否自动上传到服务器（默认True）
    
    【支持的文件格式】
    - PPT: .pptx, .ppt
    - Word: .docx, .doc  
    - PDF: .pdf
    
    【返回值】
    - 成功时返回提取结果和Excel文件信息
    - 如果auto_upload=True，还会返回公网下载链接
    - 失败时返回错误信息
    
    【使用示例】
    >>> extract_tables_from_document(
    ...     document_url="https://example.com/presentation.pptx",
    ...     output_filename="extracted_tables.xlsx"
    ... )
    
    【注意事项】
    - 需要安装相应的文档处理库（python-pptx, python-docx, pdfplumber等）
    - 大文件处理可能需要较长时间
    - PDF表格提取效果取决于PDF的格式和结构
    """
    try:
        # 验证URL格式
        if not (document_url.startswith("http://") or document_url.startswith("https://")):
            return "Error: 请输入有效的http/https链接"
        
        # 导入文档提取模块
        from excel_mcp.document_extractor import extract_tables_from_document_url
        
        # 提取表格
        result = extract_tables_from_document_url(document_url, output_filename)
        
        if not result['success']:
            return f"Error: {result['message']}"
        
        # 如果启用自动上传
        if auto_upload and result.get('output_file'):
            try:
                # 生成上传文件名
                uploaded_filename = f"extracted_{uuid.uuid4().hex}.xlsx"
                uploaded_path = os.path.join(tempfile.gettempdir(), uploaded_filename)
                
                # 复制文件到临时目录
                import shutil
                shutil.copy(result['output_file'], uploaded_path)
                
                # 上传到服务器
                remote_path = f"/root/files/{uploaded_filename}"
                transport = paramiko.Transport(("8.156.74.79", 22))
                transport.connect(username="root", password="zfsZBC123.")
                sftp = paramiko.SFTPClient.from_transport(transport)
                
                if sftp is not None:
                    sftp.put(uploaded_path, remote_path)
                    sftp.close()
                if transport is not None:
                    transport.close()
                
                # 生成下载链接
                download_url = f"http://8.156.74.79:8001/{uploaded_filename}"
                
                # 更新结果
                result['download_url'] = download_url
                result['uploaded_filename'] = uploaded_filename
                
            except Exception as upload_error:
                logger.error(f"上传文件失败: {upload_error}")
                result['upload_error'] = str(upload_error)
        
        # 格式化返回结果
        response = {
            'success': True,
            'message': result['message'],
            'total_tables': result['total_tables'],
            'tables_info': result.get('tables_info', []),
            'output_file': result.get('output_file', ''),
            'download_url': result.get('download_url', ''),
            'uploaded_filename': result.get('uploaded_filename', '')
        }
        
        import json
        return json.dumps(response, indent=2, default=str, ensure_ascii=False)
        
    except Exception as e:
        logger.error(f"提取文档表格失败: {e}")
        return f"Error: 提取文档表格失败 - {str(e)}"

@mcp.tool()
def preview_document_tables(
    document_url: str,
    max_tables: int = 3,
    max_rows_per_table: int = 10
) -> str:
    """
    预览文档中的表格结构，不进行完整提取。
    
    【功能说明】
    - 快速预览文档中的表格数量和基本结构
    - 返回每个表格的前几行数据作为预览
    - 适用于在完整提取前了解文档内容
    
    【参数说明】
    - document_url: 文档的URL地址
    - max_tables: 最大预览表格数量（默认3个）
    - max_rows_per_table: 每个表格最大预览行数（默认10行）
    
    【返回值】
    - 文档基本信息
    - 表格预览数据
    - 建议的提取策略
    """
    try:
        # 验证URL格式
        if not (document_url.startswith("http://") or document_url.startswith("https://")):
            return "Error: 请输入有效的http/https链接"
        
        # 导入文档提取模块
        from excel_mcp.document_extractor import DocumentExtractor
        
        extractor = DocumentExtractor()
        
        # 下载文件
        temp_file = extractor.download_file(document_url)
        
        try:
            # 确定文件类型
            file_type = os.path.splitext(temp_file)[1].lower().lstrip('.')
            
            # 提取表格
            tables = extractor.extract_tables_from_document(temp_file, file_type)
            
            if not tables:
                return "未在文档中找到表格"
            
            # 限制预览的表格数量
            preview_tables = tables[:max_tables]
            
            # 构建预览结果
            preview_result = {
                'document_url': document_url,
                'file_type': file_type,
                'total_tables_found': len(tables),
                'preview_tables_count': len(preview_tables),
                'tables_preview': []
            }
            
            for i, table_info in enumerate(preview_tables, 1):
                table_data = table_info['data']
                
                # 限制预览行数
                preview_rows = table_data[:max_rows_per_table]
                
                table_preview = {
                    'table_index': i,
                    'table_info': {
                        'total_rows': len(table_data),
                        'total_columns': len(table_data[0]) if table_data else 0,
                        'preview_rows': len(preview_rows)
                    },
                    'preview_data': preview_rows,
                    'location': table_info.get('slide', table_info.get('page', table_info.get('table', 'unknown')))
                }
                
                preview_result['tables_preview'].append(table_preview)
            
            # 添加建议
            if len(tables) > max_tables:
                preview_result['suggestion'] = f"文档中共有{len(tables)}个表格，建议使用extract_tables_from_document进行完整提取"
            else:
                preview_result['suggestion'] = "可以使用extract_tables_from_document进行完整提取"
            
            import json
            return json.dumps(preview_result, indent=2, default=str, ensure_ascii=False)
            
        finally:
            # 清理临时文件
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
                    
    except Exception as e:
        logger.error(f"预览文档表格失败: {e}")
        return f"Error: 预览文档表格失败 - {str(e)}"

async def run_sse():
    """Run Excel MCP server in SSE mode."""
    # Assign value to EXCEL_FILES_PATH in SSE mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with SSE transport (files directory: {EXCEL_FILES_PATH})")
        await mcp.run_sse_async()
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

async def run_streamable_http():
    """Run Excel MCP server in streamable HTTP mode."""
    # Assign value to EXCEL_FILES_PATH in streamable HTTP mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with streamable HTTP transport (files directory: {EXCEL_FILES_PATH})")
        await mcp.run_streamable_http_async()
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_stdio():
    """Run Excel MCP server in stdio mode."""
    # No need to assign EXCEL_FILES_PATH in stdio mode
    
    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")